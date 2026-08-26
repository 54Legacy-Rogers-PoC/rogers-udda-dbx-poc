"""Shared workbook/context helpers for object-access template scripts."""

from __future__ import annotations

import importlib
import os
import tempfile
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook, load_workbook


def normalize(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def header_key(value: str) -> str:
    return " ".join(normalize(value).replace("\n", " ").split()).lower()


def normalize_environment(value: str) -> str:
    mapped = {
        "PROD": "PRD",
        "PRODUCTION": "PRD",
        "PRD": "PRD",
        "QA": "QA",
        "DEV": "DEV",
    }
    upper = value.upper()
    return mapped.get(upper, upper)


def merge_context(primary: dict[str, str], fallback: dict[str, str]) -> dict[str, str]:
    merged = dict(fallback)
    for key, value in primary.items():
        if normalize(value):
            merged[key] = value
    return merged


def load_workbook_compat(template_file: Path) -> tuple[Any, str | None]:
    # First try openpyxl regardless of file extension. This handles files whose
    # extension is .xls but content is actually OOXML (.xlsx).
    try:
        return load_workbook(filename=template_file, data_only=True), None
    except Exception as xlsx_exc:  # pylint: disable=broad-except
        if template_file.suffix.lower() != ".xls":
            raise ValueError(f"Template file unreadable as xlsx: {xlsx_exc}") from xlsx_exc

    try:
        xlrd = importlib.import_module("xlrd")
    except ModuleNotFoundError as exc:
        raise SystemExit("xlrd is required for .xls support. Install with: pip install xlrd")

    try:
        xls_book = xlrd.open_workbook(str(template_file))
    except Exception as xls_exc:  # pylint: disable=broad-except
        raise ValueError(f"Template file unreadable as xls: {xls_exc}") from xls_exc

    xlsx_book = Workbook()
    xlsx_book.remove(xlsx_book.active)
    for sheet_name in xls_book.sheet_names():
        xls_sheet = xls_book.sheet_by_name(sheet_name)
        xlsx_sheet = xlsx_book.create_sheet(title=sheet_name[:31] or "Sheet")
        for row_idx in range(xls_sheet.nrows):
            for col_idx in range(xls_sheet.ncols):
                xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=xls_sheet.cell_value(row_idx, col_idx))

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as handle:
        tmp_path = handle.name

    xlsx_book.save(tmp_path)
    return load_workbook(filename=tmp_path, data_only=True), tmp_path


def _is_placeholder(value: str) -> bool:
    v = value.strip().lower()
    return v.startswith("<enter") or v in {"na", "n/a", "-"}


def _find_requestor_entry_col(sheet: Any) -> int | None:
    for row in sheet.iter_rows(min_row=1, max_row=30, values_only=True):
        for idx, cell in enumerate([normalize(v) for v in row]):
            if "requestor entry info" in header_key(cell):
                return idx
    return None


def _value_right_of_label(
    sheet: Any,
    label_matchers: list[str],
    requestor_col: int | None = None,
    prefer_email: bool = False,
) -> str:
    def _value_ok(candidate: str) -> bool:
        if not candidate or _is_placeholder(candidate):
            return False
        return (not prefer_email) or ("@" in candidate)

    def _value_from_entry_col(row_values: list[str]) -> str:
        if requestor_col is None or requestor_col >= len(row_values):
            return ""
        entry_val = row_values[requestor_col]
        return entry_val if _value_ok(entry_val) else ""

    def _value_right_of_index(row_values: list[str], start_idx: int) -> str:
        for nxt in row_values[start_idx + 1 :]:
            if _value_ok(nxt):
                return nxt
        return ""

    for row in sheet.iter_rows(values_only=True):
        row_values = [normalize(v) for v in row]
        for idx, cell in enumerate(row_values):
            cell_key = header_key(cell)
            if any(m in cell_key for m in label_matchers):
                entry_value = _value_from_entry_col(row_values)
                if entry_value:
                    return entry_value
                right_value = _value_right_of_index(row_values, idx)
                if right_value:
                    return right_value
    return ""


def read_workbook_context(
    template_file: Path,
    normalize_activity: Callable[[str], str],
    *,
    swallow_errors: bool,
) -> dict[str, str]:
    def _empty_context() -> dict[str, str]:
        return {
            "environment": "",
            "activity_type": "",
            "ad_group_name": "",
            "service_account_name": "",
        }

    def _extract_sheet_context(sheet: Any, requestor_col: int) -> dict[str, str]:
        return {
            "environment": _value_right_of_label(sheet, ["environment"], requestor_col=requestor_col),
            "activity_type": _value_right_of_label(sheet, ["request type", "activity"], requestor_col=requestor_col),
            "ad_group_name": _value_right_of_label(
                sheet,
                ["dtb ad group name", "ad group name"],
                requestor_col=requestor_col,
                prefer_email=True,
            ),
            "service_account_name": _value_right_of_label(
                sheet,
                ["service account name"],
                requestor_col=requestor_col,
                prefer_email=True,
            ),
        }

    def _merge_found_context(context_values: dict[str, str], found_values: dict[str, str]) -> None:
        for key, value in found_values.items():
            if not context_values[key] and value:
                context_values[key] = value

    def _context_complete(context_values: dict[str, str]) -> bool:
        return bool(
            context_values["environment"]
            and context_values["activity_type"]
            and (context_values["ad_group_name"] or context_values["service_account_name"])
        )

    def _principal_fields(context_values: dict[str, str]) -> tuple[str, str]:
        if context_values["ad_group_name"]:
            return "ad_group", context_values["ad_group_name"]
        if context_values["service_account_name"]:
            return "service_account", context_values["service_account_name"]
        return "", ""

    try:
        workbook, tmp_path = load_workbook_compat(template_file)
    except Exception:  # pylint: disable=broad-except
        if swallow_errors:
            return {}
        raise

    try:
        context_values = _empty_context()

        for sheet in workbook.worksheets:
            requestor_col = _find_requestor_entry_col(sheet)
            if requestor_col is None:
                continue

            _merge_found_context(context_values, _extract_sheet_context(sheet, requestor_col))
            if _context_complete(context_values):
                break

        access_for, principal_name = _principal_fields(context_values)

        return {
            "environment": normalize_environment(context_values["environment"]),
            "access_for": access_for,
            "activity_type": normalize_activity(context_values["activity_type"]),
            "ad_group_name": context_values["ad_group_name"],
            "service_account_name": context_values["service_account_name"],
            "principal_name": principal_name,
        }
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass