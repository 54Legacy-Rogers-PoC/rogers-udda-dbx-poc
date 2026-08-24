"""Generate request YAML from an object access template workbook.

The workbook may contain two sheets:
1) Request info sheet (key/value style form fields)
2) Object list sheet (table-style object rows)

CLI values override sheet-derived values.
"""

from __future__ import annotations

import argparse
from datetime import date, datetime
import os
import re
import sys
import tempfile
from pathlib import Path
from typing import Any

try:
	import yaml
except ImportError as exc:  # pragma: no cover - runtime dependency
	raise SystemExit("PyYAML is required. Install with: pip install pyyaml") from exc

try:
	from openpyxl import Workbook, load_workbook
except ImportError as exc:  # pragma: no cover - runtime dependency
	raise SystemExit("openpyxl is required. Install with: pip install openpyxl") from exc

try:
	import xlrd
except ImportError:
	xlrd = None


REQUEST_ID_REGEX = re.compile(r"^RITM[0-9]+$")

HEADER_ALIASES = {
	"Activity": {"Activity", "activity"},
	"Environment": {"Environment", "ENV", "Env", "environment"},
	"Access_For": {"Access_For", "Access_for", "access_for"},
	"Principal_Name": {"Principal_Name", "principal_name"},
	"Justification": {"Justification", "justification"},
}

REQUEST_INFO_FIELD_ALIASES = {
	"request_date": {"Request Date"},
	"requestor_name": {"Requestor Name"},
	"requestor_department_scoa": {"Requestor Department SCOA"},
	"requestor_department": {"Requestor Department"},
	"request_type": {"Request Type"},
	"environment": {"Environment"},
	"ad_group_name": {
		"DTB AD Group Name (only for individual direct access to databricks)",
		"DTB AD Group Name",
	},
	"service_account_name": {
		"Service Account Name (only for non-individual direct access to databricks)",
		"Service Account Name",
	},
	"justification": {
		"Detailed justification for an data access",
		"Detailed justification for data access",
		"Justification",
	},
	"additional_information": {
		"Approval evidence from data owners/data access approver",
		"Group Owner Director/Sr Manager Approval Message",
	},
	"group_owner_approval_message": {"Group Owner Director/Sr Manager Approval Message"},
	"group_owner_approval_date": {"Group Owner Approval Date"},
	"data_access_approval_evidence": {"Approval evidence from data owners/data access approver"},
}

REQUEST_TYPE_TO_ACTIVITY = {
	"ADD OBJECTS": "ADD",
	"ADD": "ADD",
	"REMOVE OBJECTS": "REMOVE",
	"REMOVE": "REMOVE",
	"REVOKE": "REVOKE",
}

ENVIRONMENT_ALIASES = {
	"PROD": "PRD",
	"PRODUCTION": "PRD",
	"PRD": "PRD",
	"QA": "QA",
	"DEV": "DEV",
}


def _normalize(value: Any) -> str:
	if value is None:
		return ""
	if isinstance(value, datetime):
		return value.date().isoformat()
	if isinstance(value, date):
		return value.isoformat()
	return str(value).strip()


def _is_placeholder(value: str) -> bool:
	lower = value.strip().lower()
	return lower.startswith("<enter ") or lower.startswith("<provide ") or lower.startswith("<activity ")


def _canonical_headers(raw_headers: list[str]) -> list[str]:
	headers: list[str] = []
	for header in raw_headers:
		if header == "":
			headers.append("")
			continue

		mapped = None
		for canonical, aliases in HEADER_ALIASES.items():
			if header in aliases:
				mapped = canonical
				break
		headers.append(mapped or header)
	return headers


def _first_data_row(workbook: Any, sheet_name: str) -> dict[str, str]:
	if sheet_name not in workbook.sheetnames:
		raise ValueError(f"Worksheet not found: {sheet_name}")

	sheet = workbook[sheet_name]
	header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
	raw_headers = [_normalize(v) for v in (header_cells or [])]
	headers = _canonical_headers(raw_headers)

	for row_values in sheet.iter_rows(min_row=2, values_only=True):
		row_dict = {headers[i]: row_values[i] if i < len(row_values) else None for i in range(len(headers))}
		if all(_normalize(v) == "" for v in row_dict.values()):
			continue
		return {k: _normalize(v) for k, v in row_dict.items() if k}

	raise ValueError("No non-empty rows found in worksheet")


def _map_request_info_field(label: str) -> str:
	for canonical_name, aliases in REQUEST_INFO_FIELD_ALIASES.items():
		if label in aliases:
			return canonical_name
	return ""


def _read_request_info(workbook: Any, request_sheet_name: str) -> dict[str, str]:
	if request_sheet_name not in workbook.sheetnames:
		raise ValueError(f"Worksheet not found: {request_sheet_name}")

	sheet = workbook[request_sheet_name]
	parsed: dict[str, str] = {}

	for row in sheet.iter_rows(values_only=True):
		if not row:
			continue

		label = _normalize(row[0] if len(row) > 0 else "")
		entry = _normalize(row[2] if len(row) > 2 else "")

		if label == "":
			continue

		canonical_field = _map_request_info_field(label)
		if canonical_field == "":
			continue

		if entry == "" or _is_placeholder(entry):
			continue

		parsed[canonical_field] = entry

	request_type = _normalize(parsed.get("request_type")).upper()
	if request_type in REQUEST_TYPE_TO_ACTIVITY:
		parsed["activity_type"] = REQUEST_TYPE_TO_ACTIVITY[request_type]

	environment = _normalize(parsed.get("environment")).upper()
	if environment in ENVIRONMENT_ALIASES:
		parsed["environment"] = ENVIRONMENT_ALIASES[environment]

	if _normalize(parsed.get("ad_group_name")) != "":
		parsed["access_for"] = "ad_group"
	elif _normalize(parsed.get("service_account_name")) != "":
		parsed["access_for"] = "service_account"

	request_date = _normalize(parsed.get("request_date"))
	if len(request_date) >= 10:
		parsed["request_date"] = request_date[:10]

	return parsed


def _validate_request_id(request_id: str) -> None:
	if not REQUEST_ID_REGEX.match(request_id):
		raise ValueError("request_id must match pattern RITM<digits>")


def _load_yaml_template(template_file: Path) -> dict[str, Any]:
	if not template_file.exists():
		return {}

	with template_file.open("r", encoding="utf-8") as handle:
		loaded = yaml.safe_load(handle)

	if loaded is None:
		return {}
	if not isinstance(loaded, dict):
		raise ValueError("YAML template root must be a mapping")
	return loaded


def _load_workbook_compat(template_file: Path) -> tuple[Any, str | None]:
	if template_file.suffix.lower() == ".xlsx":
		return load_workbook(filename=template_file, data_only=True), None

	if template_file.suffix.lower() != ".xls":
		raise ValueError("Template file must be .xlsx or .xls")

	if xlrd is None:
		raise SystemExit("xlrd is required for .xls support. Install with: pip install xlrd")

	xls_book = xlrd.open_workbook(str(template_file))
	xlsx_book = Workbook()
	xlsx_book.remove(xlsx_book.active)

	for sheet_name in xls_book.sheet_names():
		xls_sheet = xls_book.sheet_by_name(sheet_name)
		xlsx_sheet = xlsx_book.create_sheet(title=sheet_name[:31] or "Sheet")
		for row_idx in range(xls_sheet.nrows):
			for col_idx in range(xls_sheet.ncols):
				xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=xls_sheet.cell_value(row_idx, col_idx))

	with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as handle:
		tmp_path = handle.name

	xlsx_book.save(tmp_path)
	return load_workbook(filename=tmp_path, data_only=True), tmp_path


def _build_payload(
	request_id: str,
	template_file_name: str,
	request_info: dict[str, str],
	first_row: dict[str, str],
	platform: str,
	request_type: str,
	environment: str,
	activity_type: str,
	access_for: str,
	ad_group_name: str,
	service_account_name: str,
	justification: str,
	additional_information: str,
) -> dict[str, str]:

	info_environment = _normalize(request_info.get("environment")).upper()
	info_activity = _normalize(request_info.get("activity_type")).upper()
	info_access_for = _normalize(request_info.get("access_for")).lower()
	info_request_date = _normalize(request_info.get("request_date"))
	info_requestor_name = _normalize(request_info.get("requestor_name"))
	info_requestor_department_scoa = _normalize(request_info.get("requestor_department_scoa"))
	info_requestor_department = _normalize(request_info.get("requestor_department"))
	info_ad_group_name = _normalize(request_info.get("ad_group_name"))
	info_service_account_name = _normalize(request_info.get("service_account_name"))
	info_justification = _normalize(request_info.get("justification"))
	info_additional_information = _normalize(request_info.get("additional_information"))
	info_group_owner_approval_message = _normalize(request_info.get("group_owner_approval_message")) or info_additional_information
	info_group_owner_approval_date = _normalize(request_info.get("group_owner_approval_date"))
	info_data_access_approval_evidence = _normalize(request_info.get("data_access_approval_evidence")) or info_additional_information

	row_environment = _normalize(first_row.get("Environment")).upper()
	row_activity = _normalize(first_row.get("Activity")).upper()
	row_access_for = _normalize(first_row.get("Access_For")).lower()
	row_principal = _normalize(first_row.get("Principal_Name"))
	row_justification = _normalize(first_row.get("Justification"))

	effective_environment = environment.upper() if environment else (info_environment or row_environment)
	effective_activity = activity_type.upper() if activity_type else (info_activity or row_activity)
	effective_access_for = access_for.lower() if access_for else (info_access_for or row_access_for)
	effective_principal = (
		ad_group_name
		or service_account_name
		or info_ad_group_name
		or info_service_account_name
		or row_principal
	)
	effective_justification = justification or info_justification or row_justification
	effective_additional_information = additional_information or info_additional_information

	if effective_access_for not in {"ad_group", "service_account"}:
		raise ValueError("access_for must be one of: ad_group, service_account")

	if effective_access_for == "ad_group":
		final_ad_group_name = effective_principal
		final_service_account_name = ""
	else:
		final_ad_group_name = ""
		final_service_account_name = effective_principal

	if effective_environment == "":
		raise ValueError("environment is required (CLI or template row)")
	if effective_activity == "":
		raise ValueError("activity_type is required (CLI or template row)")
	if effective_principal == "":
		raise ValueError("principal name is required (CLI or template row)")
	if effective_justification == "":
		raise ValueError("justification is required (CLI or template row)")

	return {
		"request_id": request_id,
		"request_date": info_request_date,
		"requestor_name": info_requestor_name,
		"requestor_department_scoa": info_requestor_department_scoa,
		"requestor_department": info_requestor_department,
		"platform": platform,
		"request_type": request_type,
		"environment": effective_environment,
		"activity_type": effective_activity,
		"access_for": effective_access_for,
		"ad_group_name": final_ad_group_name,
		"ad_group_description": "",
		"ad_group_owner_name": "",
		"ad_group_owner_email": "",
		"service_account_name": final_service_account_name,
		"service_account_owner_name": "",
		"service_account_owner_email": "",
		"template_file": template_file_name,
		"justification": effective_justification,
		"additional_information": effective_additional_information,
		"group_owner_approval_message": info_group_owner_approval_message,
		"group_owner_approval_date": info_group_owner_approval_date,
		"data_access_approval_evidence": info_data_access_approval_evidence,
	}


def parse_args() -> argparse.Namespace:
	parser = argparse.ArgumentParser(description="Generate request YAML from ObjectAccess template")
	parser.add_argument("--request-id", required=True, help="Request id, e.g. RITM123456")
	parser.add_argument("--template-file", required=True, help="Path to ObjectAccessTemplate.xlsx")
	parser.add_argument("--output-file", help="Output YAML path. Default: uda/requests/object-access/<request_id>.yaml")
	parser.add_argument(
		"--yaml-template-file",
		default="",
		help="Optional request YAML template to pre-populate fields",
	)
	parser.add_argument("--sheet-name", default="ObjectAccess", help="Object list worksheet name")
	parser.add_argument("--request-sheet-name", default="", help="Request info worksheet name (default: first sheet)")
	parser.add_argument("--platform", default="databricks", help="platform field")
	parser.add_argument("--request-type", default="object_access", help="request_type field")
	parser.add_argument("--environment", default="", help="Override environment")
	parser.add_argument("--activity-type", default="", help="Override activity_type")
	parser.add_argument("--access-for", default="", help="Override access_for: ad_group|service_account")
	parser.add_argument("--ad-group-name", default="", help="Override ad_group_name")
	parser.add_argument("--service-account-name", default="", help="Override service_account_name")
	parser.add_argument("--justification", default="", help="Override justification")
	parser.add_argument("--additional-information", default="", help="additional_information field")
	return parser.parse_args()


def main() -> int:
	args = parse_args()
	request_id = args.request_id.strip()

	try:
		_validate_request_id(request_id)
		template_file = Path(args.template_file).resolve()
		if not template_file.exists():
			raise FileNotFoundError(f"Template file not found: {template_file}")

		workbook, tmp_path = _load_workbook_compat(template_file)
		request_sheet_name = args.request_sheet_name or workbook.sheetnames[0]
		try:
			request_info = _read_request_info(workbook=workbook, request_sheet_name=request_sheet_name)

			first_row = _first_data_row(workbook=workbook, sheet_name=args.sheet_name)
			payload = _build_payload(
				request_id=request_id,
				template_file_name=template_file.name,
				request_info=request_info,
				first_row=first_row,
				platform=args.platform,
				request_type=args.request_type,
				environment=args.environment,
				activity_type=args.activity_type,
				access_for=args.access_for,
				ad_group_name=args.ad_group_name,
				service_account_name=args.service_account_name,
				justification=args.justification,
				additional_information=args.additional_information,
			)
		finally:
			if tmp_path:
				try:
					os.unlink(tmp_path)
				except OSError:
					pass

		if args.yaml_template_file:
			yaml_template_file = Path(args.yaml_template_file).resolve()
		else:
			yaml_template_file = (
				Path(__file__).resolve().parents[2]
				/ "templates"
				/ "object-access"
				/ "object-access-yaml-template.yaml"
			)

		template_payload = _load_yaml_template(yaml_template_file)
		payload = {**template_payload, **payload}

		if args.output_file:
			output_file = Path(args.output_file).resolve()
		else:
			output_file = (
				Path(__file__).resolve().parents[2]
				/ "requests"
				/ "object-access"
				/ f"{request_id}.yaml"
			)

		output_file.parent.mkdir(parents=True, exist_ok=True)
		output_file.write_text(yaml.safe_dump(payload, sort_keys=False), encoding="utf-8")
		print(f"Request YAML written: {output_file}")
		return 0

	except Exception as exc:  # pylint: disable=broad-except
		print(f"Generation failed: {exc}")
		return 1


if __name__ == "__main__":
	sys.exit(main())
