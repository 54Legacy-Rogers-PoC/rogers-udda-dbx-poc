"""Parse Databricks object access template into normalized records.

This script converts ObjectAccess worksheet rows into a consistent JSON structure
used by tfvars generation and Terraform provisioning stages.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

try:
	import yaml
except ImportError as exc:  # pragma: no cover - runtime dependency
	raise SystemExit("PyYAML is required. Install with: pip install pyyaml") from exc

try:
	from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - runtime dependency
	raise SystemExit("openpyxl is required. Install with: pip install openpyxl") from exc


REQUIRED_COLUMNS = [
	"Environment",
	"Access_For",
	"Principal_Name",
	"Object_Type",
	"Catalog",
	"Schema",
	"Object_Name",
]

COMPACT_REQUIRED_COLUMNS = ["Activity", "Catalog", "Schema", "Object_Name"]

HEADER_ALIASES = {
	"Record_ID": {"Record_ID", "record_id"},
	"Activity": {"Activity", "activity"},
	"Environment": {"Environment", "ENV", "Env", "environment"},
	"Access_For": {"Access_For", "Access_for", "access_for"},
	"Principal_Name": {"Principal_Name", "principal_name"},
	"Object_Type": {"Object_Type", "object_type"},
	"Catalog": {"Catalog", "catalog", "UC Catalog"},
	"Schema": {
		"Schema",
		"schema",
		"EDL Schema Name/Folder Name (vw_*) / Sandbox Name (vw_slfsrv_*) For non-EDL / <schema name>",
	},
	"Object_Name": {"Object_Name", "Object", "View", "view", "object_name", "Object Name"},
	"Folder_Path": {"Folder_Path", "folder_path"},
	"Privilege": {"Privilege", "privilege"},
	"Justification": {"Justification", "justification"},
	"Additional_Information": {"Additional_Information", "additional_information"},
}


def _normalize(value: Any) -> str:
	if value is None:
		return ""
	return str(value).strip()


def _header_key(value: str) -> str:
	return " ".join(_normalize(value).replace("\n", " ").split()).lower()


def _normalize_environment(value: str) -> str:
	mapped = {
		"PROD": "PRD",
		"PRODUCTION": "PRD",
		"PRD": "PRD",
		"QA": "QA",
		"DEV": "DEV",
	}
	upper = value.upper()
	return mapped.get(upper, upper)


def _normalize_activity(value: str) -> str:
	mapped = {
		"ADD": "ADD",
		"ADD OBJECT": "ADD",
		"ADD OBJECTS": "ADD",
		"REMOVE": "REMOVE",
		"REMOVE OBJECT": "REMOVE",
		"REMOVE OBJECTS": "REMOVE",
		"REVOKE": "REMOVE",
	}
	upper = value.upper()
	return mapped.get(upper, upper)


def _is_blank(value: Any) -> bool:
	return _normalize(value) == ""


def _default_privilege(object_type: str) -> str:
	defaults = {
		"CATALOG": "USE_CATALOG",
		"SCHEMA": "USE_SCHEMA",
		"VIEW": "SELECT",
		"FOLDER": "READ",
	}
	return defaults.get(object_type.upper(), "")


def _canonical_headers(raw_headers: list[str]) -> list[str]:
	canonical_headers: list[str] = []
	header_alias_keys: dict[str, set[str]] = {
		canonical: {_header_key(alias) for alias in aliases}
		for canonical, aliases in HEADER_ALIASES.items()
	}

	for header in raw_headers:
		if header == "":
			canonical_headers.append("")
			continue

		mapped = None
		normalized_header = _header_key(header)
		for canonical, aliases in HEADER_ALIASES.items():
			if normalized_header in header_alias_keys[canonical]:
				mapped = canonical
				break

		canonical_headers.append(mapped or header)

	return canonical_headers


def _read_request_context(request_file: Path | None) -> dict[str, str]:
	if request_file is None:
		return {}

	with request_file.open("r", encoding="utf-8") as handle:
		payload = yaml.safe_load(handle)

	if not isinstance(payload, dict):
		return {}

	access_for = _normalize(payload.get("access_for")).lower()
	principal_name = ""
	if access_for == "ad_group":
		principal_name = _normalize(payload.get("ad_group_name"))
	elif access_for == "service_account":
		principal_name = _normalize(payload.get("service_account_name"))

	return {
		"request_id": _normalize(payload.get("request_id")),
		"environment": _normalize(payload.get("environment")).upper(),
		"access_for": access_for,
		"activity_type": _normalize(payload.get("activity_type")).upper(),
		"principal_name": principal_name,
		"justification": _normalize(payload.get("justification")),
	}


def _load_rows(template_file: Path, sheet_name: str) -> tuple[list[dict[str, Any]], list[str]]:
	if not template_file.exists():
		raise FileNotFoundError(f"Template file not found: {template_file}")

	workbook = load_workbook(filename=template_file, data_only=True)
	if sheet_name not in workbook.sheetnames:
		raise ValueError(f"Worksheet not found: {sheet_name}")

	sheet = workbook[sheet_name]
	header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
	raw_headers = [_normalize(v) for v in (header_cells or [])]
	if _header_key(raw_headers[0] if raw_headers else "") == "information entered by requestor":
		header_cells = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True), None)
		raw_headers = [_normalize(v) for v in (header_cells or [])]
		min_row = 3
	else:
		min_row = 2
	headers = _canonical_headers(raw_headers)

	is_compact_format = not ({"Access_For", "Principal_Name", "Object_Type"} & set(headers))
	required_columns = COMPACT_REQUIRED_COLUMNS if is_compact_format else REQUIRED_COLUMNS
	missing = [column for column in required_columns if column not in headers]
	if missing:
		raise ValueError(f"Missing required columns: {', '.join(missing)}")

	rows: list[dict[str, Any]] = []
	for row_idx, row_values in enumerate(sheet.iter_rows(min_row=min_row, values_only=True), start=min_row):
		row_dict = {headers[i]: row_values[i] if i < len(row_values) else None for i in range(len(headers))}
		if all(_is_blank(v) for v in row_dict.values()):
			continue
		row_dict["_row"] = row_idx
		rows.append(row_dict)

	return rows, headers


def _normalize_row(row: dict[str, Any], request_context: dict[str, str]) -> dict[str, Any]:
	row_activity = _normalize(row.get("Activity")).upper()
	row_environment = _normalize(row.get("Environment")).upper()
	row_access_for = _normalize(row.get("Access_For")).lower()
	row_principal = _normalize(row.get("Principal_Name"))

	activity = _normalize_activity(row_activity or request_context.get("activity_type", ""))
	environment = _normalize_environment(row_environment or request_context.get("environment", ""))
	access_for = row_access_for or request_context.get("access_for", "")
	principal_name = row_principal or request_context.get("principal_name", "")
	catalog = _normalize(row.get("Catalog"))
	schema = _normalize(row.get("Schema"))
	object_name = _normalize(row.get("Object_Name"))
	folder_path = _normalize(row.get("Folder_Path"))

	object_type = _normalize(row.get("Object_Type")).upper()
	if object_type == "":
		if folder_path != "":
			object_type = "FOLDER"
		elif object_name != "":
			object_type = "VIEW"
		elif schema != "":
			object_type = "SCHEMA"
		elif catalog != "":
			object_type = "CATALOG"

	privilege = _normalize(row.get("Privilege")).upper() or _default_privilege(object_type)
	record_id = _normalize(row.get("Record_ID")) or f"ROW-{row['_row']:04d}"
	justification = _normalize(row.get("Justification")) or request_context.get("justification", "")

	return {
		"row_number": row["_row"],
		"record_id": record_id,
		"activity": activity,
		"environment": environment,
		"access_for": access_for,
		"principal_name": principal_name,
		"object_type": object_type,
		"catalog": catalog,
		"schema": schema,
		"object_name": object_name,
		"folder_path": folder_path,
		"privilege": privilege,
		"justification": justification,
		"additional_information": _normalize(row.get("Additional_Information")),
	}


def parse_template_file(
	template_file: Path,
	request_file: Path | None = None,
	sheet_name: str = "ObjectAccess",
) -> dict[str, Any]:
	request_context = _read_request_context(request_file)
	rows, headers = _load_rows(template_file=template_file, sheet_name=sheet_name)
	records = [_normalize_row(row, request_context) for row in rows]

	return {
		"request_id": request_context.get("request_id", ""),
		"sheet_name": sheet_name,
		"header_columns": headers,
		"record_count": len(records),
		"records": records,
	}


def parse_args() -> argparse.Namespace:
	parser = argparse.ArgumentParser(description="Parse object access template workbook")
	parser.add_argument("--template-file", required=True, help="Path to ObjectAccessTemplate.xlsx")
	parser.add_argument("--request-file", help="Path to request YAML file")
	parser.add_argument("--sheet-name", default="ObjectAccess", help="Worksheet name")
	parser.add_argument("--output-file", help="Output JSON file path")
	return parser.parse_args()


def main() -> int:
	args = parse_args()
	template_file = Path(args.template_file).resolve()
	request_file = Path(args.request_file).resolve() if args.request_file else None

	try:
		parsed = parse_template_file(
			template_file=template_file,
			request_file=request_file,
			sheet_name=args.sheet_name,
		)
	except Exception as exc:  # pylint: disable=broad-except
		print(f"Parse failed: {exc}")
		return 1

	json_payload = json.dumps(parsed, indent=2)
	if args.output_file:
		output_file = Path(args.output_file).resolve()
		output_file.parent.mkdir(parents=True, exist_ok=True)
		output_file.write_text(json_payload + "\n", encoding="utf-8")
		print(f"Parsed output written: {output_file}")
	else:
		print(json_payload)

	return 0


if __name__ == "__main__":
	sys.exit(main())
