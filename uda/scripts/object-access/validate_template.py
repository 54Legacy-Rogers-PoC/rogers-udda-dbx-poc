"""Validate Databricks object access Excel templates.

This script validates worksheet structure and row-level object access rules before
Terraform plan/apply execution.
"""

from __future__ import annotations

import argparse
import importlib
import os
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

try:
	import yaml
except ImportError as exc:  # pragma: no cover - runtime dependency
	raise SystemExit("PyYAML is required. Install with: pip install pyyaml") from exc

try:
	from openpyxl import Workbook, load_workbook
except ImportError as exc:  # pragma: no cover - runtime dependency
	raise SystemExit("openpyxl is required. Install with: pip install openpyxl") from exc

REQUIRED_COLUMNS = [
	"Environment",
	"Access_For",
	"Principal_Name",
	"Object_Type",
	"Catalog",
	"Schema",
	"Object_Name",
]

COMPACT_REQUIRED_COLUMNS = ["Activity", "Catalog", "Schema", "Object_Name"]

OPTIONAL_COLUMNS = ["Record_ID", "Activity", "Folder_Path", "Privilege", "Justification", "Additional_Information"]

HEADER_ALIASES = {
	"Record_ID": {"Record_ID", "record_id"},
	"Activity": {"Activity", "activity"},
	"Environment": {"Environment", "ENV", "Env", "environment"},
	"Access_For": {"Access_For", "Access_for", "access_for"},
	"Principal_Name": {"Principal_Name", "principal_name"},
	"Object_Type": {"Object_Type", "object_type"},
	"Catalog": {"Catalog", "catalog", "UC Catalog"},
	"Schema": {
		"Schema",
		"schema",
		"EDL Schema Name/Folder Name (vw_*) / Sandbox Name (vw_slfsrv_*) For non-EDL / <schema name>",
	},
	"Object_Name": {"Object_Name", "Object", "View", "view", "object_name", "Object Name"},
	"Folder_Path": {"Folder_Path", "folder_path"},
	"Privilege": {"Privilege", "privilege"},
	"Justification": {"Justification", "justification"},
	"Additional_Information": {"Additional_Information", "additional_information"},
}

ALLOWED_ACTIVITY = {"ADD", "REMOVE", "REVOKE"}
ALLOWED_ENV = {"DEV", "QA", "PRD"}
ALLOWED_ACCESS_FOR = {"ad_group", "service_account"}
ALLOWED_OBJECT_TYPES = {"CATALOG", "SCHEMA", "VIEW", "FOLDER"}

ALLOWED_PRIVILEGES = {
	"CATALOG": {"USE_CATALOG"},
	"SCHEMA": {"USE_SCHEMA", "CREATE_TABLE", "CREATE_VIEW"},
	"VIEW": {"SELECT"},
	"FOLDER": {"READ", "WRITE", "READ_WRITE"},
}

DUPLICATE_FIELDS = [
	"Activity",
	"Environment",
	"Access_For",
	"Principal_Name",
	"Object_Type",
	"Catalog",
	"Schema",
	"Object_Name",
	"Folder_Path",
	"Privilege",
]


@dataclass
class ValidationError:
	code: str
	row: int
	field: str
	message: str


def _normalize(value: Any) -> str:
	if value is None:
		return ""
	return str(value).strip()


def _header_key(value: str) -> str:
	return " ".join(_normalize(value).replace("\n", " ").split()).lower()


def _normalize_environment(value: str) -> str:
	mapped = {
		"PROD": "PRD",
		"PRODUCTION": "PRD",
		"PRD": "PRD",
		"QA": "QA",
		"DEV": "DEV",
	}
	upper = value.upper()
	return mapped.get(upper, upper)


def _normalize_principal_for_compare(value: str) -> str:
	v = _normalize(value)
	if "@" in v:
		return v.lower()
	return v


def _normalize_activity(value: str) -> str:
	mapped = {
		"ADD": "ADD",
		"ADD OBJECT": "ADD",
		"ADD OBJECTS": "ADD",
		"REMOVE": "REMOVE",
		"REMOVE OBJECT": "REMOVE",
		"REMOVE OBJECTS": "REMOVE",
		"REVOKE": "REVOKE",
	}
	upper = value.upper()
	return mapped.get(upper, upper)


def _is_blank(value: Any) -> bool:
	return _normalize(value) == ""


def _default_privilege(object_type: str) -> str:
	defaults = {
		"CATALOG": "USE_CATALOG",
		"SCHEMA": "USE_SCHEMA",
		"VIEW": "SELECT"
		
	}
	return defaults.get(object_type.upper(), "")


def _effective_privilege(record: dict[str, Any]) -> str:
	object_type = _normalize(record.get("Object_Type")).upper()
	row_privilege = _normalize(record.get("Privilege")).upper()
	return row_privilege or _default_privilege(object_type)


def _canonical_headers(raw_headers: list[str]) -> tuple[list[str], list[str]]:
	canonical_headers: list[str] = []
	unrecognized: list[str] = []
	header_alias_keys: dict[str, set[str]] = {
		canonical: {_header_key(alias) for alias in aliases}
		for canonical, aliases in HEADER_ALIASES.items()
	}

	for header in raw_headers:
		if header == "":
			canonical_headers.append("")
			continue

		mapped = None
		normalized_header = _header_key(header)
		for canonical, aliases in HEADER_ALIASES.items():
			if normalized_header in header_alias_keys[canonical]:
				mapped = canonical
				break

		if mapped is None:
			canonical_headers.append(header)
			unrecognized.append(header)
		else:
			canonical_headers.append(mapped)

	return canonical_headers, unrecognized


def _read_request_context(request_file: Path | None) -> dict[str, str]:
	if request_file is None:
		return {}

	with request_file.open("r", encoding="utf-8") as handle:
		payload = yaml.safe_load(handle)

	if not isinstance(payload, dict):
		return {}

	return {
		"environment": _normalize(payload.get("environment")),
		"access_for": _normalize(payload.get("access_for")),
		"activity_type": _normalize(payload.get("activity_type")),
		"ad_group_name": _normalize(payload.get("ad_group_name")),
		"service_account_name": _normalize(payload.get("service_account_name")),
	}


def _is_placeholder(value: str) -> bool:
	v = value.strip().lower()
	return v.startswith("<enter") or v in {"na", "n/a", "-"}


def _find_requestor_entry_col(sheet: Any) -> int | None:
	for row in sheet.iter_rows(min_row=1, max_row=30, values_only=True):
		for idx, cell in enumerate([_normalize(v) for v in row]):
			if "requestor entry info" in _header_key(cell):
				return idx
	return None


def _value_right_of_label(sheet: Any, label_matchers: list[str], requestor_col: int | None = None, prefer_email: bool = False) -> str:
	for row in sheet.iter_rows(values_only=True):
		row_values = [_normalize(v) for v in row]
		for idx, cell in enumerate(row_values):
			cell_key = _header_key(cell)
			if any(m in cell_key for m in label_matchers):
				if requestor_col is not None and requestor_col < len(row_values):
					entry_val = row_values[requestor_col]
					if entry_val and not _is_placeholder(entry_val):
						if not prefer_email or "@" in entry_val:
							return entry_val
				for nxt in row_values[idx + 1 :]:
					if nxt and not _is_placeholder(nxt):
						if prefer_email and "@" not in nxt:
							continue
						return nxt
	return ""


def _extract_sheet_context(sheet: Any, requestor_col: int) -> dict[str, str]:
	return {
		"environment": _value_right_of_label(sheet, ["environment"], requestor_col=requestor_col),
		"activity_type": _value_right_of_label(sheet, ["request type", "activity"], requestor_col=requestor_col),
		"ad_group_name": _value_right_of_label(
			sheet,
			["dtb ad group name", "ad group name"],
			requestor_col=requestor_col,
			prefer_email=True,
		),
		"service_account_name": _value_right_of_label(
			sheet,
			["service account name"],
			requestor_col=requestor_col,
			prefer_email=True,
		),
	}


def _merge_context_values(current: dict[str, str], found: dict[str, str]) -> None:
	for key, value in found.items():
		if not current[key] and value:
			current[key] = value


def _has_complete_context(context_values: dict[str, str]) -> bool:
	return bool(
		context_values["environment"]
		and context_values["activity_type"]
		and (context_values["ad_group_name"] or context_values["service_account_name"])
	)


def _read_workbook_context(template_file: Path) -> dict[str, str]:
	try:
		workbook, tmp_path = _load_workbook_compat(template_file)
	except Exception:  # pylint: disable=broad-except
		return {}

	try:
		context_values = {
			"environment": "",
			"activity_type": "",
			"ad_group_name": "",
			"service_account_name": "",
		}

		for sheet in workbook.worksheets:
			requestor_col = _find_requestor_entry_col(sheet)
			if requestor_col is None:
				continue
			found = _extract_sheet_context(sheet, requestor_col)
			_merge_context_values(context_values, found)

			if _has_complete_context(context_values):
				break

		access_for = ""
		if context_values["ad_group_name"]:
			access_for = "ad_group"
		elif context_values["service_account_name"]:
			access_for = "service_account"

		return {
			"environment": _normalize_environment(context_values["environment"]),
			"access_for": access_for,
			"activity_type": _normalize_activity(context_values["activity_type"]),
			"ad_group_name": context_values["ad_group_name"],
			"service_account_name": context_values["service_account_name"],
		}
	finally:
		if tmp_path:
			try:
				os.unlink(tmp_path)
			except OSError:
				pass


def _merge_context(primary: dict[str, str], fallback: dict[str, str]) -> dict[str, str]:
	merged = dict(fallback)
	for key, value in primary.items():
		if _normalize(value):
			merged[key] = value
	return merged


def _load_workbook_compat(template_file: Path) -> tuple[Any, str | None]:
	# First try openpyxl regardless of file extension. This handles files whose
	# extension is .xls but content is actually OOXML (.xlsx).
	try:
		return load_workbook(filename=template_file, data_only=True), None
	except Exception as xlsx_exc:  # pylint: disable=broad-except
		if template_file.suffix.lower() != ".xls":
			raise ValueError(f"Template file unreadable as xlsx: {xlsx_exc}") from xlsx_exc

	try:
		xlrd = importlib.import_module("xlrd")
	except ModuleNotFoundError as exc:
		raise SystemExit("xlrd is required for .xls support. Install with: pip install xlrd")

	try:
		xls_book = xlrd.open_workbook(str(template_file))
	except Exception as xls_exc:  # pylint: disable=broad-except
		raise ValueError(f"Template file unreadable as xls: {xls_exc}") from xls_exc
	xlsx_book = Workbook()
	xlsx_book.remove(xlsx_book.active)

	for sheet_name in xls_book.sheet_names():
		xls_sheet = xls_book.sheet_by_name(sheet_name)
		xlsx_sheet = xlsx_book.create_sheet(title=sheet_name[:31] or "Sheet")
		for row_idx in range(xls_sheet.nrows):
			for col_idx in range(xls_sheet.ncols):
				xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=xls_sheet.cell_value(row_idx, col_idx))

	with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as handle:
		tmp_path = handle.name

	xlsx_book.save(tmp_path)
	return load_workbook(filename=tmp_path, data_only=True), tmp_path


def _load_rows(template_file: Path, sheet_name: str) -> tuple[list[dict[str, Any]], list[ValidationError]]:
	if not template_file.exists():
		return [], [
			ValidationError(
				code="TPL-001",
				row=0,
				field="template_file",
				message=f"Template file not found: {template_file}",
			)
		]

	try:
		workbook, tmp_path = _load_workbook_compat(template_file)
	except Exception as exc:  # pylint: disable=broad-except
		return [], [
			ValidationError(
				code="TPL-001",
				row=0,
				field="template_file",
				message=f"Template file unreadable: {exc}",
			)
		]

	try:
		if sheet_name not in workbook.sheetnames:
			return [], [
				ValidationError(
					code="TPL-002",
					row=0,
					field="worksheet",
					message=f"Worksheet not found: {sheet_name}",
				)
			]

		sheet = workbook[sheet_name]
		header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
		raw_headers = [_normalize(v) for v in (header_cells or [])]
		if _header_key(raw_headers[0] if raw_headers else "") == "information entered by requestor":
			header_cells = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True), None)
			raw_headers = [_normalize(v) for v in (header_cells or [])]
			min_row = 3
		else:
			min_row = 2
		headers, _ = _canonical_headers(raw_headers)

		is_compact_format = not ({"Access_For", "Principal_Name", "Object_Type"} & set(headers))
		required_columns = COMPACT_REQUIRED_COLUMNS if is_compact_format else REQUIRED_COLUMNS
		missing = [c for c in required_columns if c not in headers]
		if missing:
			return [], [
				ValidationError(
					code="TPL-003",
					row=1,
					field="columns",
					message=f"Missing required columns: {', '.join(missing)}",
				)
			]

		records: list[dict[str, Any]] = []
		for row_idx, row_values in enumerate(sheet.iter_rows(min_row=min_row, values_only=True), start=min_row):
			row_dict = {headers[i]: row_values[i] if i < len(row_values) else None for i in range(len(headers))}
			# Ignore fully empty rows.
			if all(_is_blank(v) for v in row_dict.values()):
				continue
			row_dict["_row"] = row_idx
			records.append(row_dict)

		return records, []
	finally:
		if tmp_path:
			try:
				os.unlink(tmp_path)
			except OSError:
				pass


def _hydrate_record(record: dict[str, Any], request_context: dict[str, str]) -> dict[str, Any]:
	hydrated = dict(record)

	row_activity = _normalize(hydrated.get("Activity"))
	hydrated["Activity"] = _normalize_activity(row_activity or request_context.get("activity_type", ""))

	row_environment = _normalize(hydrated.get("Environment"))
	hydrated["Environment"] = _normalize_environment(row_environment or request_context.get("environment", ""))

	row_access_for = _normalize(hydrated.get("Access_For")).lower()
	hydrated["Access_For"] = row_access_for or request_context.get("access_for", "")

	row_principal = _normalize(hydrated.get("Principal_Name"))
	expected_principal = ""
	if hydrated["Access_For"] == "ad_group":
		expected_principal = request_context.get("ad_group_name", "")
	elif hydrated["Access_For"] == "service_account":
		expected_principal = request_context.get("service_account_name", "")
	hydrated["Principal_Name"] = row_principal or expected_principal

	catalog = _normalize(hydrated.get("Catalog"))
	schema = _normalize(hydrated.get("Schema"))
	object_name = _normalize(hydrated.get("Object_Name"))
	folder_path = _normalize(hydrated.get("Folder_Path"))

	object_type = _normalize(hydrated.get("Object_Type")).upper()
	if object_type == "":
		if folder_path != "":
			object_type = "FOLDER"
		elif object_name != "":
			object_type = "VIEW"
		elif schema != "":
			object_type = "SCHEMA"
		elif catalog != "":
			object_type = "CATALOG"
	hydrated["Object_Type"] = object_type

	return hydrated


def _validate_row_required(record: dict[str, Any], request_context: dict[str, str]) -> list[ValidationError]:
	errors: list[ValidationError] = []
	row_no = record["_row"]

	mandatory_fields = ["Environment", "Access_For", "Principal_Name", "Object_Type"]

	for field in mandatory_fields:
		if _is_blank(record.get(field)):
			errors.append(
				ValidationError(
					code="TPL-010",
					row=row_no,
					field=field,
					message=f"Mandatory field is blank: {field}",
				)
			)

	activity = _normalize(record.get("Activity"))
	if activity == "" and request_context.get("activity_type", "") == "":
		errors.append(
			ValidationError(
				code="TPL-010",
				row=row_no,
				field="Activity",
				message="Activity is blank and no request-level activity_type was provided",
			)
		)

	if _effective_privilege(record) == "":
		errors.append(
			ValidationError(
				code="TPL-010",
				row=row_no,
				field="Privilege",
				message="Privilege is missing and no default can be inferred from Object_Type",
			)
		)

	return errors


def _validate_row_enums(record: dict[str, Any], request_context: dict[str, str]) -> list[ValidationError]:
	errors: list[ValidationError] = []
	row_no = record["_row"]

	activity = _normalize(record.get("Activity"))
	if activity == "":
		activity = request_context.get("activity_type", "")
	if activity and activity.upper() not in ALLOWED_ACTIVITY:
		errors.append(
			ValidationError(
				code="TPL-004",
				row=row_no,
				field="Activity",
				message="Invalid Activity value",
			)
		)

	environment = _normalize(record.get("Environment")).upper()
	if environment and environment not in ALLOWED_ENV:
		errors.append(
			ValidationError(
				code="TPL-004",
				row=row_no,
				field="Environment",
				message="Invalid Environment value",
			)
		)

	access_for = _normalize(record.get("Access_For")).lower()
	if access_for and access_for not in ALLOWED_ACCESS_FOR:
		errors.append(
			ValidationError(
				code="TPL-004",
				row=row_no,
				field="Access_For",
				message="Invalid Access_For value",
			)
		)

	object_type = _normalize(record.get("Object_Type")).upper()
	if object_type and object_type not in ALLOWED_OBJECT_TYPES:
		errors.append(
			ValidationError(
				code="TPL-004",
				row=row_no,
				field="Object_Type",
				message="Invalid Object_Type value",
			)
		)

	return errors


def _validate_conditional_fields(record: dict[str, Any]) -> list[ValidationError]:
	errors: list[ValidationError] = []
	row_no = record["_row"]
	object_type = _normalize(record.get("Object_Type")).upper()

	catalog = _normalize(record.get("Catalog"))
	schema = _normalize(record.get("Schema"))
	object_name = _normalize(record.get("Object_Name"))
	folder_path = _normalize(record.get("Folder_Path"))

	if object_type == "CATALOG":
		if catalog == "" or schema or object_name or folder_path:
			errors.append(
				ValidationError(
					code="TPL-005",
					row=row_no,
					field="Object_Type",
					message="CATALOG requires Catalog and disallows Schema/Object_Name/Folder_Path",
				)
			)

	if object_type == "SCHEMA":
		if catalog == "" or schema == "" or object_name or folder_path:
			errors.append(
				ValidationError(
					code="TPL-005",
					row=row_no,
					field="Object_Type",
					message="SCHEMA requires Catalog and Schema and disallows Object_Name/Folder_Path",
				)
			)

	if object_type == "VIEW":
		if catalog == "" or schema == "" or object_name == "" or folder_path:
			errors.append(
				ValidationError(
					code="TPL-005",
					row=row_no,
					field="Object_Type",
					message="VIEW requires Catalog/Schema/Object_Name and disallows Folder_Path",
				)
			)

	if object_type == "FOLDER":
		if folder_path == "" or catalog or schema or object_name:
			errors.append(
				ValidationError(
					code="TPL-005",
					row=row_no,
					field="Object_Type",
					message="FOLDER requires Folder_Path and disallows Catalog/Schema/Object_Name",
				)
			)

	return errors


def _validate_privilege(record: dict[str, Any]) -> list[ValidationError]:
	row_no = record["_row"]
	object_type = _normalize(record.get("Object_Type")).upper()
	privilege = _effective_privilege(record)

	if object_type not in ALLOWED_PRIVILEGES:
		return []

	if privilege not in ALLOWED_PRIVILEGES[object_type]:
		return [
			ValidationError(
				code="TPL-007",
				row=row_no,
				field="Privilege",
				message=(
					f"Invalid privilege for {object_type}. "
					f"Allowed: {sorted(ALLOWED_PRIVILEGES[object_type])}"
				),
			)
		]

	return []


def _validate_cross_field_consistency(record: dict[str, Any], request_context: dict[str, str]) -> list[ValidationError]:
	errors: list[ValidationError] = []
	row_no = record["_row"]

	row_env = _normalize(record.get("Environment")).upper()
	row_access_for = _normalize(record.get("Access_For")).lower()
	row_principal = _normalize(record.get("Principal_Name"))

	expected_env = request_context.get("environment", "")
	expected_access_for = request_context.get("access_for", "")

	if expected_env and row_env and row_env != expected_env:
		errors.append(
			ValidationError(
				code="TPL-008",
				row=row_no,
				field="Environment",
				message=f"Row environment {row_env} does not match request environment {expected_env}",
			)
		)

	if expected_access_for and row_access_for and row_access_for != expected_access_for:
		errors.append(
			ValidationError(
				code="TPL-008",
				row=row_no,
				field="Access_For",
				message=(
					f"Row access_for {row_access_for} does not match "
					f"request access_for {expected_access_for}"
				),
			)
		)

	expected_principal = ""
	if expected_access_for == "ad_group":
		expected_principal = request_context.get("ad_group_name", "")
	elif expected_access_for == "service_account":
		expected_principal = request_context.get("service_account_name", "")

	if expected_principal and row_principal and _normalize_principal_for_compare(row_principal) != _normalize_principal_for_compare(expected_principal):
		errors.append(
			ValidationError(
				code="TPL-008",
				row=row_no,
				field="Principal_Name",
				message=(
					f"Row principal {row_principal} does not match request principal "
					f"{expected_principal}"
				),
			)
		)

	return errors


def _validate_duplicates(records: list[dict[str, Any]], request_context: dict[str, str]) -> list[ValidationError]:
	errors: list[ValidationError] = []
	seen: dict[tuple[str, ...], int] = {}

	for record in records:
		activity = _normalize(record.get("Activity"))
		if activity == "":
			activity = request_context.get("activity_type", "")

		key_parts = []
		for field in DUPLICATE_FIELDS:
			if field == "Activity":
				key_parts.append(activity.upper())
			elif field == "Environment":
				key_parts.append(_normalize(record.get(field)).upper())
			elif field == "Access_For":
				key_parts.append(_normalize(record.get(field)).lower())
			elif field == "Object_Type":
				key_parts.append(_normalize(record.get(field)).upper())
			elif field == "Privilege":
				key_parts.append(_effective_privilege(record))
			else:
				key_parts.append(_normalize(record.get(field)))

		key = tuple(key_parts)
		if key in seen:
			first_row = seen[key]
			errors.append(
				ValidationError(
					code="TPL-006",
					row=record["_row"],
					field="row",
					message=f"Duplicate row detected (first occurrence at row {first_row})",
				)
			)
		else:
			seen[key] = record["_row"]

	return errors


def _effective_activity(record: dict[str, Any], request_context: dict[str, str]) -> str:
	activity = _normalize(record.get("Activity"))
	if activity == "":
		activity = request_context.get("activity_type", "")
	effective = activity.upper()
	if effective == "REVOKE":
		return "REMOVE"
	return effective


def _validate_add_prerequisites(records: list[dict[str, Any]], request_context: dict[str, str]) -> list[ValidationError]:
	errors: list[ValidationError] = []

	catalog_keys: set[tuple[str, str, str, str]] = set()
	schema_keys: set[tuple[str, str, str, str, str]] = set()

	view_rows: list[tuple[int, tuple[str, str, str, str], tuple[str, str, str, str, str]]] = []
	schema_rows: list[tuple[int, tuple[str, str, str, str]]] = []
	catalog_privileges: dict[tuple[str, str, str, str], set[str]] = {}
	schema_privileges: dict[tuple[str, str, str, str, str], set[str]] = {}

	for record in records:
		if _effective_activity(record, request_context) != "ADD":
			continue

		env = _normalize(record.get("Environment")).upper()
		access_for = _normalize(record.get("Access_For")).lower()
		principal = _normalize(record.get("Principal_Name")).lower()
		catalog = _normalize(record.get("Catalog"))
		schema = _normalize(record.get("Schema"))
		object_type = _normalize(record.get("Object_Type")).upper()

		catalog_key = (env, access_for, principal, catalog)
		schema_key = (env, access_for, principal, catalog, schema)

		if object_type == "CATALOG":
			catalog_keys.add(catalog_key)
			catalog_privileges.setdefault(catalog_key, set()).add(_effective_privilege(record))
		elif object_type == "SCHEMA":
			schema_keys.add(schema_key)
			schema_privileges.setdefault(schema_key, set()).add(_effective_privilege(record))
			schema_rows.append((record["_row"], catalog_key))
		elif object_type == "VIEW":
			view_rows.append((record["_row"], catalog_key, schema_key))

	# Dependency checks are needed only when VIEW grants are present.
	if not view_rows:
		return []

	# Compact intake templates may provide only VIEW rows and rely on existing
	# catalog/schema grants. In that mode, skip prerequisite enforcement.
	if view_rows and not schema_rows and not catalog_keys:
		return []

	for row_no, catalog_key in schema_rows:
		if catalog_key not in catalog_keys:
			errors.append(
				ValidationError(
					code="TPL-011",
					row=row_no,
					field="Object_Type",
					message="SCHEMA ADD requires matching CATALOG ADD row for the same Environment/Access_For/Principal/Catalog",
				)
			)
		if "USE_CATALOG" not in catalog_privileges.get(catalog_key, set()):
			errors.append(
				ValidationError(
					code="TPL-011",
					row=row_no,
					field="Privilege",
					message="SCHEMA ADD requires matching CATALOG ADD privilege USE_CATALOG for the same Environment/Access_For/Principal/Catalog",
				)
			)

	for row_no, catalog_key, schema_key in view_rows:
		if schema_key not in schema_keys:
			errors.append(
				ValidationError(
					code="TPL-011",
					row=row_no,
					field="Object_Type",
					message="VIEW ADD requires matching SCHEMA ADD row for the same Environment/Access_For/Principal/Catalog/Schema",
				)
			)
		if "USE_SCHEMA" not in schema_privileges.get(schema_key, set()):
			errors.append(
				ValidationError(
					code="TPL-011",
					row=row_no,
					field="Privilege",
					message="VIEW ADD requires matching SCHEMA ADD privilege USE_SCHEMA for the same Environment/Access_For/Principal/Catalog/Schema",
				)
			)
		if catalog_key not in catalog_keys:
			errors.append(
				ValidationError(
					code="TPL-011",
					row=row_no,
					field="Object_Type",
					message="VIEW ADD requires matching CATALOG ADD row for the same Environment/Access_For/Principal/Catalog",
				)
			)
		if "USE_CATALOG" not in catalog_privileges.get(catalog_key, set()):
			errors.append(
				ValidationError(
					code="TPL-011",
					row=row_no,
					field="Privilege",
					message="VIEW ADD requires matching CATALOG ADD privilege USE_CATALOG for the same Environment/Access_For/Principal/Catalog",
				)
			)

	return errors


def validate_template_file(
	template_file: Path,
	request_file: Path | None = None,
	sheet_name: str = "ObjectAccess",
	max_rows: int = 1000,
) -> list[ValidationError]:
	request_context = _read_request_context(request_file)
	workbook_context = _read_workbook_context(template_file)
	request_context = _merge_context(request_context, workbook_context)

	rows, load_errors = _load_rows(template_file, sheet_name)
	if load_errors:
		return load_errors

	hydrated_rows = [_hydrate_record(record, request_context) for record in rows]

	errors: list[ValidationError] = []

	if len(hydrated_rows) > max_rows:
		errors.append(
			ValidationError(
				code="TPL-009",
				row=0,
				field="row_count",
				message=f"Template contains {len(hydrated_rows)} rows; max allowed is {max_rows}",
			)
		)

	for record in hydrated_rows:
		errors.extend(_validate_row_required(record, request_context))
		errors.extend(_validate_row_enums(record, request_context))
		errors.extend(_validate_conditional_fields(record))
		errors.extend(_validate_privilege(record))
		errors.extend(_validate_cross_field_consistency(record, request_context))

	errors.extend(_validate_add_prerequisites(hydrated_rows, request_context))
	errors.extend(_validate_duplicates(hydrated_rows, request_context))
	return errors


def parse_args() -> argparse.Namespace:
	parser = argparse.ArgumentParser(description="Validate object access template workbook")
	parser.add_argument("--template-file", required=True, help="Path to ObjectAccessTemplate.xlsx")
	parser.add_argument("--request-file", help="Path to request YAML file for cross-field checks")
	parser.add_argument("--sheet-name", default="ObjectAccess", help="Worksheet name")
	parser.add_argument("--max-rows", default=1000, type=int, help="Maximum allowed data rows")
	return parser.parse_args()


def main() -> int:
	args = parse_args()
	template_file = Path(args.template_file).resolve()
	request_file = Path(args.request_file).resolve() if args.request_file else None

	errors = validate_template_file(
		template_file=template_file,
		request_file=request_file,
		sheet_name=args.sheet_name,
		max_rows=args.max_rows,
	)

	if errors:
		print("Template validation failed.")
		for err in errors:
			print(f"- [{err.code}] row={err.row} field={err.field}: {err.message}")
		return 1

	print("Template validation successful.")
	return 0


if __name__ == "__main__":
	sys.exit(main())
