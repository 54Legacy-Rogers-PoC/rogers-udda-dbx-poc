import argparse
import csv
import json
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
import yaml


class ValidationError(Exception):
    pass


@dataclass
class RequestContext:
    request_id: str
    platform: str
    request_type: str
    environment: str
    activity_type: str
    access_for: str
    principal_name: str
    template_file: str
    justification: str
    additional_information: str


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def load_yaml(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as file:
        return yaml.safe_load(file) or {}


def normalize_field(name: str) -> str:
    return name.strip().lower().replace(" ", "_")


def map_environment(raw_environment: str, config: dict[str, Any]) -> str:
    mapping = config.get("environment_mapping", {})
    if raw_environment in mapping:
        return mapping[raw_environment]

    raise ValidationError(f"Unsupported environment value: {raw_environment}")


def parse_request(request: dict[str, Any], config: dict[str, Any]) -> RequestContext:
    required = [
        "request_id",
        "platform",
        "request_type",
        "environment",
        "activity_type",
        "access_for",
        "template_file",
        "justification",
    ]
    for field in required:
        if not request.get(field):
            raise ValidationError(f"Missing required field in request YAML: {field}")

    access_for = str(request["access_for"]).strip().lower()
    if access_for not in {"ad_group", "service_account"}:
        raise ValidationError("access_for must be either ad_group or service_account")

    principal_name_field = "ad_group_name" if access_for == "ad_group" else "service_account_name"
    principal_name = request.get(principal_name_field)
    if not principal_name:
        raise ValidationError(f"Missing required field for {access_for}: {principal_name_field}")

    return RequestContext(
        request_id=str(request["request_id"]).strip(),
        platform=str(request["platform"]).strip().lower(),
        request_type=str(request["request_type"]).strip().lower(),
        environment=map_environment(str(request["environment"]).strip(), config),
        activity_type=str(request["activity_type"]).strip().upper(),
        access_for=access_for,
        principal_name=str(principal_name).strip(),
        template_file=str(request["template_file"]).strip(),
        justification=str(request["justification"]).strip(),
        additional_information=str(request.get("additional_information", "")).strip(),
    )


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        return [dict(row) for row in reader]


def read_xlsx_rows(path: Path, worksheet_name: str) -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if worksheet_name not in workbook.sheetnames:
        raise ValidationError(f"Worksheet '{worksheet_name}' not found in template")

    worksheet = workbook[worksheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    normalized_headers = [normalize_field(h) for h in headers]

    result: list[dict[str, str]] = []
    for row in rows[1:]:
        if row is None:
            continue

        row_data: dict[str, str] = {}
        has_value = False
        for idx, value in enumerate(row):
            key = normalized_headers[idx] if idx < len(normalized_headers) else f"extra_{idx}"
            parsed = "" if value is None else str(value).strip()
            row_data[key] = parsed
            has_value = has_value or bool(parsed)

        if has_value:
            result.append(row_data)

    return result


def load_template_rows(template_path: Path, template_config: dict[str, Any]) -> list[dict[str, str]]:
    if not template_path.exists():
        raise ValidationError(f"Template file not found: {template_path}")

    suffix = template_path.suffix.lower()
    if suffix == ".csv":
        csv_rows = read_csv_rows(template_path)
        return [{normalize_field(k): (v or "").strip() for k, v in row.items()} for row in csv_rows]

    if suffix == ".xlsx":
        return read_xlsx_rows(template_path, template_config["worksheet_name"])

    raise ValidationError("Template must be either .xlsx or .csv")


def require_columns(rows: list[dict[str, str]], required_columns: list[str]) -> None:
    if not rows:
        raise ValidationError("Template does not contain any data rows")

    available = set(rows[0].keys())
    required = {normalize_field(c) for c in required_columns}
    missing = sorted(required - available)
    if missing:
        raise ValidationError(f"Template missing required columns: {', '.join(missing)}")


def parse_privileges(raw: str) -> list[str]:
    privileges = [p.strip().upper() for p in raw.split(",") if p and p.strip()]
    if not privileges:
        raise ValidationError("Privileges column cannot be empty")
    return privileges


def validate_record(
    row: dict[str, str],
    row_number: int,
    context: RequestContext,
    template_config: dict[str, Any],
) -> dict[str, Any]:
    object_type = row.get("object_type", "").strip().lower()
    if object_type not in template_config["allowed_object_types"]:
        raise ValidationError(f"Row {row_number}: invalid object type '{object_type}'")

    row_activity = row.get("activity", "").strip().upper() or context.activity_type
    if row_activity not in template_config["allowed_activities"]:
        raise ValidationError(f"Row {row_number}: invalid activity '{row_activity}'")

    privileges = parse_privileges(row.get("privileges", ""))

    if object_type == "folder":
        allowed = set(template_config["folder_permission_levels"])
        invalid_folder_privs = [p for p in privileges if p not in allowed]
        if invalid_folder_privs:
            raise ValidationError(
                f"Row {row_number}: invalid folder permission level(s): {', '.join(invalid_folder_privs)}"
            )

    record = {
        "row_id": f"{context.request_id}-{row_number:05d}",
        "activity": row_activity,
        "object_type": object_type,
        "principal_type": context.access_for,
        "principal_name": context.principal_name,
        "catalog_name": row.get("catalog_name", "").strip() or None,
        "schema_name": row.get("schema_name", "").strip() or None,
        "object_name": row.get("object_name", "").strip() or None,
        "folder_path": row.get("folder_path", "").strip() or None,
        "privileges": privileges,
        "justification": context.justification,
    }

    if object_type == "catalog" and not record["catalog_name"]:
        raise ValidationError(f"Row {row_number}: catalog_name is required for catalog access")
    if object_type == "schema" and (not record["catalog_name"] or not record["schema_name"]):
        raise ValidationError(f"Row {row_number}: catalog_name and schema_name are required for schema access")
    if object_type == "view" and (
        not record["catalog_name"] or not record["schema_name"] or not record["object_name"]
    ):
        raise ValidationError(
            f"Row {row_number}: catalog_name, schema_name, and object_name are required for view access"
        )
    if object_type == "folder" and not record["folder_path"]:
        raise ValidationError(f"Row {row_number}: folder_path is required for folder access")

    return record


def dedupe_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[str] = set()
    result: list[dict[str, Any]] = []

    for record in records:
        duplicate_key = json.dumps(
            {
                "activity": record["activity"],
                "object_type": record["object_type"],
                "principal_type": record["principal_type"],
                "principal_name": record["principal_name"],
                "catalog_name": record["catalog_name"],
                "schema_name": record["schema_name"],
                "object_name": record["object_name"],
                "folder_path": record["folder_path"],
                "privileges": sorted(record["privileges"]),
            },
            sort_keys=True,
        )

        if duplicate_key in seen:
            raise ValidationError(
                f"Duplicate template record detected for principal {record['principal_name']} row {record['row_id']}"
            )

        seen.add(duplicate_key)
        result.append(record)

    return result


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as file:
        json.dump(payload, file, indent=2)


def main() -> None:
    parser = argparse.ArgumentParser(description="Process UDA Databricks object access requests")
    parser.add_argument("--request-file", required=True)
    parser.add_argument("--config-file", required=True)
    parser.add_argument("--output-dir", required=True)
    args = parser.parse_args()

    started_at = utc_now()

    request_file = Path(args.request_file)
    config_file = Path(args.config_file)
    output_dir = Path(args.output_dir)

    if not request_file.exists():
        raise ValidationError(f"Request file not found: {request_file}")
    if not config_file.exists():
        raise ValidationError(f"Config file not found: {config_file}")

    config = load_yaml(config_file)
    template_config = config.get("template", {})

    request_data = load_yaml(request_file)
    context = parse_request(request_data, config)

    if context.platform != "databricks":
        raise ValidationError("platform must be databricks")
    if context.request_type != "object_access":
        raise ValidationError("request_type must be object_access")

    attachments_root = Path("uda/attachments/object-access")
    template_path = attachments_root / context.template_file
    rows = load_template_rows(template_path, template_config)

    require_columns(rows, template_config["required_columns"])

    max_records = int(template_config.get("max_records", 1000))
    if len(rows) > max_records:
        raise ValidationError(f"Template exceeds max_records limit of {max_records}")

    validated_records = [
        validate_record(row=row, row_number=index + 2, context=context, template_config=template_config)
        for index, row in enumerate(rows)
    ]
    validated_records = dedupe_records(validated_records)

    terraform_vars = {
        "request_id": context.request_id,
        "environment": context.environment,
        "object_access_records": validated_records,
    }

    terraform_vars_file = output_dir / "terraform.auto.tfvars.json"
    write_json(terraform_vars_file, terraform_vars)

    metadata = {
        "request_id": context.request_id,
        "environment": context.environment,
        "principal_type": context.access_for,
        "principal_name": context.principal_name,
        "activity_type": context.activity_type,
        "object_count": len(validated_records),
        "terraform_vars_file": str(terraform_vars_file),
    }
    write_json(output_dir / "request_metadata.json", metadata)

    execution_log = {
        "request_id": context.request_id,
        "principal_type": context.access_for,
        "principal_name": context.principal_name,
        "environment": context.environment,
        "activity_type": context.activity_type,
        "object_count": len(validated_records),
        "workflow_start_time": started_at,
        "workflow_end_time": utc_now(),
        "execution_status": "SUCCESS",
    }
    write_json(output_dir / "execution_log.json", execution_log)


if __name__ == "__main__":
    try:
        main()
    except ValidationError as exc:
        print(f"ERROR: {exc}")
        raise SystemExit(1)
